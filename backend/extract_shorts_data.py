"""
쇼츠 콘텐츠 기획 시스템을 위한 Excel 데이터 추출 스크립트
"""
import openpyxl
import json
from pathlib import Path

def extract_sub_niche_data(ws):
    """서브니치 플레이북 데이터 추출"""
    data = {
        "니치_세분화_방법": [],
        "단계별_질문": []
    }
    
    # 니치 세분화 방법 (행 2-7)
    for row_idx in range(2, 8):
        category = ws.cell(row_idx, 1).value
        examples = ws.cell(row_idx, 2).value
        if category and examples:
            data["니치_세분화_방법"].append({
                "카테고리": category,
                "예시": examples
            })
    
    # 단계별 질문 (행 10부터)
    for row_idx in range(11, ws.max_row + 1):
        step = ws.cell(row_idx, 1).value
        question = ws.cell(row_idx, 2).value
        if step and question:
            data["단계별_질문"].append({
                "단계": step,
                "질문": question
            })
    
    return data

def extract_backstory_data(ws):
    """백스토리 빌더 데이터 추출"""
    data = {
        "스토리_요소": []
    }
    
    # 스토리 요소 추출 (행 4-10)
    story_elements = [
        ("문제_또는_도전", 4),
        ("내적_갈등", 5),
        ("외적_갈등", 6),
        ("변화_이벤트", 8),
        ("영감_순간", 9),
        ("가이드_멘토", 10)
    ]
    
    for element_name, row_idx in story_elements:
        label = ws.cell(row_idx, 1).value
        example = ws.cell(row_idx, 2).value
        if label:
            data["스토리_요소"].append({
                "요소": element_name,
                "라벨": label,
                "예시": example if example else ""
            })
    
    return data

def extract_content_matrix_data(ws):
    """콘텐츠 매트릭스 데이터 추출 - 쇼츠용"""
    data = {
        "콘텐츠_구조": [],
        "쇼츠_최적화_팁": {
            "타겟": "쇼츠 시청자 (짧은 주의 집중 시간)",
            "최적_길이": "15-60초",
            "핵심_전략": [
                "첫 3초에 훅 배치",
                "빠른 컷 편집",
                "자막 필수 사용",
                "트렌드 음악 활용",
                "명확한 메시지",
                "시각적 임팩트"
            ]
        }
    }
    
    # 콘텐츠 타입 (열 B-I, 행 2-3)
    content_types = []
    for col_idx in range(2, 10):  # B-I열
        type_name = ws.cell(2, col_idx).value
        description = ws.cell(3, col_idx).value
        if type_name:
            # 쇼츠용으로 재해석
            shorts_tips = get_shorts_tips_for_type(type_name)
            content_types.append({
                "타입": type_name,
                "설명": description if description else "",
                "쇼츠_적용법": shorts_tips
            })
    
    data["콘텐츠_구조"] = content_types
    
    return data

def get_shorts_tips_for_type(content_type):
    """각 콘텐츠 타입별 쇼츠 적용 팁"""
    tips_map = {
        "Actionable": {
            "훅": "\"3가지 방법으로...\" 로 시작",
            "구조": "단계별로 빠르게 보여주기",
            "시각화": "숫자와 체크리스트 활용",
            "마무리": "\"지금 바로 시도해보세요\""
        },
        "Motivational": {
            "훅": "\"이것 때문에 포기하려 했는데...\"",
            "구조": "Before → After 변화 강조",
            "시각화": "감정 표현과 음악",
            "마무리": "\"당신도 할 수 있습니다\""
        },
        "Analytical": {
            "훅": "\"이것의 진짜 비밀은...\"",
            "구조": "문제 → 분석 → 인사이트",
            "시각화": "데이터와 그래프",
            "마무리": "핵심 포인트 요약"
        },
        "Contrarian": {
            "훅": "\"모두가 틀렸습니다\"",
            "구조": "일반 상식 → 반박 → 진실",
            "시각화": "대비 효과",
            "마무리": "\"이제 제대로 하세요\""
        },
        "Observation": {
            "훅": "\"아무도 모르는 사실\"",
            "구조": "관찰 → 발견 → 의미",
            "시각화": "실제 사례",
            "마무리": "\"주목해야 할 이유\""
        },
        "X vs. Y": {
            "훅": "\"A vs B, 진짜 차이는?\"",
            "구조": "비교 → 차이점 → 결론",
            "시각화": "분할 화면",
            "마무리": "명확한 승자 선언"
        },
        "Present / Future": {
            "훅": "\"2025년에는 이렇게 바뀝니다\"",
            "구조": "현재 → 트렌드 → 미래",
            "시각화": "타임라인",
            "마무리": "\"준비하세요\""
        },
        "Listicle": {
            "훅": "\"TOP 5 반전\"",
            "구조": "카운트다운 방식",
            "시각화": "번호와 아이콘",
            "마무리": "\"1위는?\""
        }
    }
    
    return tips_map.get(content_type, {
        "훅": "강력한 첫 문장",
        "구조": "명확한 흐름",
        "시각화": "시각적 요소 활용",
        "마무리": "행동 유도"
    })

def extract_copywriting_data(ws):
    """카피라이팅 구조 데이터 추출"""
    data = {
        "쇼츠_스크립트_구조": {
            "훅_Trailer": {
                "목적": "첫 3초 안에 시청자 붙잡기",
                "예시": []
            },
            "본론_Meat": {
                "목적": "핵심 가치 전달",
                "예시": []
            },
            "마무리_Summary_CTC": {
                "목적": "행동 유도",
                "예시": []
            }
        },
        "쇼츠_훅_템플릿": [
            "❌ [문제] → ✅ [해결책]",
            "🤔 궁금하지 않으세요?",
            "⚠️ 이거 모르면 큰일",
            "💰 돈 버는 비밀",
            "🔥 지금 핫한 이유",
            "😱 충격적인 사실",
            "🎯 단 3가지만",
            "⏰ 시간 없으면 이것만"
        ]
    }
    
    # 실제 예시 수집
    current_section = None
    for row_idx in range(2, min(50, ws.max_row + 1)):
        label = ws.cell(row_idx, 1).value
        content = ws.cell(row_idx, 2).value
        
        if label and "trailer" in str(label).lower():
            current_section = "훅_Trailer"
        elif label and "meat" in str(label).lower():
            current_section = "본론_Meat"
        elif label and content and current_section:
            data["쇼츠_스크립트_구조"][current_section]["예시"].append(content)
    
    return data

def extract_hashtag_data(ws):
    """해시태그 데이터 추출 - 쇼츠용으로 변환"""
    data = {
        "쇼츠_해시태그": {
            "필수_태그": [
                "#쇼츠", "#Shorts", "#유튜브쇼츠",
                "#short", "#shortvideo", "#shortsfeed"
            ],
            "카테고리별_태그": {}
        }
    }
    
    # 원본 해시태그를 쇼츠용으로 변환
    categories = {
        "마케팅": ["#마케팅", "#디지털마케팅", "#온라인마케팅", "#마케팅전략"],
        "창업": ["#창업", "#스타트업", "#사업", "#창업준비"],
        "성장": ["#성장", "#자기계발", "#동기부여", "#성공"],
        "투자": ["#투자", "#재테크", "#부동산", "#주식"],
        "라이프스타일": ["#일상", "#브이로그", "#루틴", "#라이프"],
        "교육": ["#교육", "#공부", "#학습", "#강의"],
        "기술": ["#기술", "#IT", "#개발", "#프로그래밍"]
    }
    
    data["쇼츠_해시태그"]["카테고리별_태그"] = categories
    
    # 원본 엑셀의 해시태그도 수집
    for row_idx in range(1, min(100, ws.max_row + 1)):
        hashtag = ws.cell(row_idx, 1).value
        if hashtag and "#" in str(hashtag):
            # LinkedIn 태그를 일반 태그로 변환
            tag_name = hashtag.split("-")[0].strip().replace("#", "")
            if tag_name and tag_name not in ["India"]:  # 불필요한 태그 제외
                for category in categories.keys():
                    if len(categories[category]) < 10:
                        categories[category].append(f"#{tag_name}")
                        break
    
    return data

def main():
    # Excel 파일 경로
    excel_path = Path("../-MAKE_A_COPY-_The_LinkedIn_Operating_System_Resources_1_.xlsx")
    
    if not excel_path.exists():
        print(f"❌ Excel 파일을 찾을 수 없습니다: {excel_path}")
        return
    
    print("📊 Excel 파일 로딩 중...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    shorts_system_data = {
        "시스템_이름": "쇼츠 콘텐츠 기획 시스템",
        "설명": "조회수 높은 쇼츠 콘텐츠를 기획하고 제작하는 올인원 솔루션",
        "버전": "1.0.0"
    }
    
    print("\n🎬 쇼츠 콘텐츠 기획 데이터 추출 중...")
    
    # 각 시트별 데이터 추출
    print("  ✓ 니치 세분화 데이터...")
    shorts_system_data["니치_전략"] = extract_sub_niche_data(wb["Sub-Niche Playbook"])
    
    print("  ✓ 스토리텔링 구조...")
    shorts_system_data["스토리_빌더"] = extract_backstory_data(wb["Better Backstory Builder"])
    
    print("  ✓ 콘텐츠 매트릭스...")
    shorts_system_data["콘텐츠_매트릭스"] = extract_content_matrix_data(wb["New & Improved Content Matrix"])
    
    print("  ✓ 카피라이팅 구조...")
    shorts_system_data["스크립트_구조"] = extract_copywriting_data(wb["Copywriting 101"])
    
    print("  ✓ 해시태그 전략...")
    shorts_system_data["해시태그_전략"] = extract_hashtag_data(wb["Top 50 LinkedIn Hashtags"])
    
    # 쇼츠 특화 추가 데이터
    shorts_system_data["쇼츠_최적화"] = {
        "플랫폼별_전략": {
            "유튜브_쇼츠": {
                "최적_길이": "15-60초",
                "비율": "9:16 세로",
                "업로드_시간": "오후 5-7시, 밤 9-11시",
                "핵심": "첫 3초가 생명"
            },
            "인스타그램_릴스": {
                "최적_길이": "15-30초",
                "비율": "9:16 세로",
                "업로드_시간": "점심시간, 저녁시간",
                "핵심": "트렌드 음악 필수"
            },
            "틱톡": {
                "최적_길이": "21-34초",
                "비율": "9:16 세로",
                "업로드_시간": "오전 6-10시, 저녁 7-11시",
                "핵심": "챌린지와 듀엣"
            }
        },
        "조회수_최적화_체크리스트": [
            "✅ 첫 3초에 강력한 훅",
            "✅ 빠른 컷 편집 (2-3초마다)",
            "✅ 명확한 자막 (가독성)",
            "✅ 트렌딩 음악 사용",
            "✅ 세로 비율 (9:16)",
            "✅ 밝고 선명한 화질",
            "✅ 시선 끄는 썸네일",
            "✅ 명확한 CTA",
            "✅ 적절한 해시태그 (3-5개)",
            "✅ 커뮤니티 참여 유도"
        ],
        "바이럴_요소": [
            "🔥 트렌드: 현재 인기 있는 주제",
            "😱 충격: 예상 밖의 반전",
            "😂 유머: 웃음 포인트",
            "💡 가치: 실용적인 정보",
            "❤️ 감동: 공감 가는 스토리",
            "🎯 타겟: 명확한 대상",
            "⚡ 속도: 빠른 전개",
            "🎨 비주얼: 눈길 끄는 화면"
        ]
    }
    
    # JSON 파일로 저장
    output_path = Path("../data/shorts_system.json")
    output_path.parent.mkdir(exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(shorts_system_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ 쇼츠 콘텐츠 기획 시스템 데이터 생성 완료!")
    print(f"📁 저장 위치: {output_path}")
    print(f"📊 데이터 크기: {output_path.stat().st_size / 1024:.1f} KB")
    
    # 요약 정보 출력
    print("\n📋 시스템 구성 요소:")
    for key in shorts_system_data.keys():
        if key not in ["시스템_이름", "설명", "버전"]:
            print(f"   • {key}")

if __name__ == "__main__":
    main()

