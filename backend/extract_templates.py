import openpyxl
import json
from pathlib import Path

def extract_excel_templates(excel_path: str):
    """Extract all template data from Excel file"""
    wb = openpyxl.load_workbook(excel_path)
    
    templates = {
        "content_matrix": {},
        "copywriting_structures": [],
        "sub_niche_playbook": {},
        "backstory_builder": {},
        "social_ecosystem": [],
        "hashtags": []
    }
    
    # Extract Content Matrix
    if "New & Improved Content Matrix" in wb.sheetnames:
        ws = wb["New & Improved Content Matrix"]
        
        # Get structure headers (row 2)
        structures = []
        for col in range(2, 12):  # Columns B to K
            cell_value = ws.cell(row=2, column=col).value
            if cell_value:
                structures.append(cell_value)
        
        # Get structure descriptions (row 3)
        descriptions = []
        for col in range(2, 12):
            cell_value = ws.cell(row=3, column=col).value
            if cell_value:
                descriptions.append(cell_value)
        
        # Get topics (column A, starting from row 8)
        topics = []
        for row in range(8, 50):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and cell_value not in ["Outcome:", "Focus:", "Structure:", "Polarity:"]:
                topics.append(cell_value)
        
        templates["content_matrix"] = {
            "structures": [
                {"name": structures[i], "description": descriptions[i]} 
                for i in range(len(structures)) if i < len(descriptions)
            ],
            "topics": topics
        }
    
    # Extract Copywriting Structure
    if "Copywriting 101" in wb.sheetnames:
        ws = wb["Copywriting 101"]
        
        structure = {
            "trailer": [],
            "meat": [],
            "summary": [],
            "ctc": []
        }
        
        current_section = None
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if row[0]:
                section_name = str(row[0]).lower()
                if "trailer" in section_name:
                    current_section = "trailer"
                elif "meat" in section_name:
                    current_section = "meat"
                elif "summary" in section_name:
                    current_section = "summary"
                elif "ctc" in section_name or "call to" in section_name:
                    current_section = "ctc"
            
            if current_section and row[1]:
                structure[current_section].append(row[1])
        
        templates["copywriting_structures"].append(structure)
    
    # Extract Sub-Niche Playbook
    if "Sub-Niche Playbook" in wb.sheetnames:
        ws = wb["Sub-Niche Playbook"]
        
        questions = []
        for row in range(10, 17):
            question = ws.cell(row=row, column=2).value
            if question:
                questions.append({
                    "step": ws.cell(row=row, column=1).value,
                    "question": question,
                    "example": ws.cell(row=row, column=3).value
                })
        
        templates["sub_niche_playbook"] = {
            "niche_methods": [
                {"name": "Price", "examples": "luxury, moderate, discount"},
                {"name": "Demographics", "examples": "Gender, age, income level, education level"},
                {"name": "Level of quality", "examples": "Premium, handmade, economical"},
                {"name": "Psychographics", "examples": "Values, interests, attitudes"},
                {"name": "Vertical/Industry", "examples": "Healthcare, finance, eComm"},
                {"name": "Geographics", "examples": "Residents of a certain country, city, or even neighborhood"}
            ],
            "questions": questions
        }
    
    # Extract Backstory Builder
    if "Better Backstory Builder" in wb.sheetnames:
        ws = wb["Better Backstory Builder"]
        
        elements = []
        for row in range(3, 10):
            element = ws.cell(row=row, column=1).value
            example = ws.cell(row=row, column=2).value
            if element:
                elements.append({
                    "element": element,
                    "example": example
                })
        
        templates["backstory_builder"] = {
            "elements": elements
        }
    
    # Extract Social Ecosystem
    if "Social Ecosystem" in wb.sheetnames:
        ws = wb["Social Ecosystem"]
        
        influencers = []
        for row in range(2, 10):
            name = ws.cell(row=row, column=1).value
            if name:
                influencers.append({
                    "name": name,
                    "platform": ws.cell(row=row, column=2).value,
                    "posting_time": ws.cell(row=row, column=3).value,
                    "url": ws.cell(row=row, column=4).value,
                    "audience": ws.cell(row=row, column=5).value
                })
        
        templates["social_ecosystem"] = influencers
    
    # Extract Hashtags
    if "Top 50 LinkedIn Hashtags" in wb.sheetnames:
        ws = wb["Top 50 LinkedIn Hashtags"]
        
        hashtags = []
        for row in range(1, 26):
            tag = ws.cell(row=row, column=1).value
            if tag:
                hashtags.append(tag)
        
        templates["hashtags"] = hashtags
    
    return templates

if __name__ == "__main__":
    excel_path = "../-MAKE_A_COPY-_The_LinkedIn_Operating_System_Resources_1_.xlsx"
    templates = extract_excel_templates(excel_path)
    
    # Save to JSON
    output_path = "../data/templates.json"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(templates, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Templates extracted successfully to {output_path}")
    print(f"📊 Found:")
    print(f"  - {len(templates['content_matrix']['structures'])} content structures")
    print(f"  - {len(templates['content_matrix']['topics'])} topics")
    print(f"  - {len(templates['sub_niche_playbook']['questions'])} niche questions")
    print(f"  - {len(templates['backstory_builder']['elements'])} backstory elements")
    print(f"  - {len(templates['hashtags'])} hashtags")

